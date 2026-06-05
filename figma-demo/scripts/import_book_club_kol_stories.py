#!/usr/bin/env python3
"""Build public/data/book_club_stories.json from KOL xlsx + live X post text (oEmbed)."""

from __future__ import annotations

import hashlib
import html
import json
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = (
    Path.home()
    / "Library/Group Containers/group.com.apple.coreservices.useractivityd/shared-pasteboard/items"
    / "516D877B-A5F6-4C20-88E2-766B6F1D5A48/Internal-币安人生书友会拟合作KOL名单.xlsx"
)
OUT_JSON = ROOT / "public/data/book_club_stories.json"
SKIP_NAMES = {"发过书相关内容", "分析型", "###", "Name", ""}

# Handles to drop from the book-club grid (user-curated + screenshot X marks).
EXCLUDE_HANDLES = {
    "0xtodd",
    "sea_bitcoin",
    "yixing_web3",
    "jason_chen998",
    "fishkiller",
    "nake13",
    "tongtongbee",
    "colinwu",
    "mindaoyang",
    "biteyecn",
    "wolfyxbt",
    "linwanwan823",
    "0xmoon",
    "zhuifei0008",
    "zhuifei008",
    "qimeiluo",
    "anndylian",
    "aster_ninja_",
    "0xcheshire",
}

MANUAL_OVERRIDES_PATH = ROOT / "scripts/manual_book_club_story_text.json"

# When the sheet has no status hyperlink, still pull live X profile bio for these accounts.
PROFILE_BIO_REFETCH_HANDLES = {"foma_bsc"}

STATUS_RE = re.compile(r"(?:x|twitter)\.com/[^/]+/status/(\d+)", re.I)
HANDLE_FROM_URL_RE = re.compile(r"(?:x|twitter)\.com/([^/]+)/status/", re.I)
OEMBED = "https://publish.twitter.com/oembed?omit_script=1&url={url}"
VXTWITTER = "https://api.vxtwitter.com/{user}/status/{tweet_id}"
FXTWITTER_USER = "https://api.fxtwitter.com/{user}"


def is_weak_post_text(text: str) -> bool:
    t = text.strip()
    if len(t) < 12:
        return True
    if re.fullmatch(r"https?://\S+", t):
        return True
    if re.fullmatch(r"https?://t\.co/\S+", t):
        return True
    return False


def fetch_vxtwitter_post(user: str, tweet_id: str) -> tuple[str | None, str | None]:
    api = VXTWITTER.format(user=urllib.parse.quote(user), tweet_id=tweet_id)
    req = urllib.request.Request(api, headers={"User-Agent": "cz-book-club-import/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=25) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError, OSError) as e:
        print(f"  vxtwitter fail: @{user}/{tweet_id} ({e})", file=sys.stderr)
        return None, None

    article = data.get("article") if isinstance(data.get("article"), dict) else None
    if article:
        title = (article.get("title") or "").strip()
        preview = (article.get("preview_text") or "").strip()
        parts = [p for p in (title, preview) if p]
        if parts:
            body = "\n\n".join(parts)
            created = (data.get("date") or "").strip()
            return body, created or None

    text = (data.get("text") or "").strip()
    if text and not is_weak_post_text(text):
        return text, (data.get("date") or "").strip() or None
    return None, None


def normalize_x_url(url: str) -> str:
    u = url.strip().replace("twitter.com", "x.com")
    if u.startswith("http://"):
        u = "https://" + u[7:]
    # drop /photo/N suffix for oEmbed
    u = re.sub(r"/photo/\d+$", "", u, flags=re.I)
    return u


def extract_paragraph_text(oembed_html: str) -> str:
    m = re.search(r"<p[^>]*>(.*?)</p>", oembed_html, re.I | re.S)
    raw = m.group(1) if m else oembed_html
    raw = re.sub(r"<a[^>]*>.*?</a>", " ", raw, flags=re.I | re.S)
    raw = re.sub(r"<[^>]+>", " ", raw)
    text = html.unescape(raw)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def fetch_x_post_text(status_url: str) -> tuple[str | None, str | None]:
    """Returns (text, created_at_hint) or (None, None) on failure."""
    url = normalize_x_url(status_url)
    api = OEMBED.format(url=urllib.parse.quote(url, safe=""))
    req = urllib.request.Request(api, headers={"User-Agent": "cz-book-club-import/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError, OSError) as e:
        print(f"  oEmbed fail: {url} ({e})", file=sys.stderr)
        return None, None

    text = extract_paragraph_text(data.get("html") or "")
    if not text:
        return None, None

    created = None
    date_m = re.search(
        r'href="https://x\.com/[^"]+/status/\d+\?[^"]*">([^<]+)</a>',
        data.get("html") or "",
        re.I,
    )
    if date_m:
        created = html.unescape(date_m.group(1)).strip()
    return text, created


def clean_sheet_reason(reason: str) -> str:
    body = re.sub(r"\s*link\s*$", "", reason, flags=re.I).strip()
    body = re.sub(r"\s+link\b", "", body, flags=re.I).strip()
    return body


def normalize_handle(handle: str) -> str:
    return handle.lstrip("@").strip().lower()


def should_exclude_handle(handle: str) -> bool:
    return normalize_handle(handle) in EXCLUDE_HANDLES


def finalize_headline_body(body: str) -> tuple[str, str]:
    """X posts have no title/body split — store full tweet text in headline only."""
    return body.strip(), ""


def load_manual_overrides() -> dict[str, dict[str, str]]:
    """Optional status-id -> {text, created_at?} for tweets that no longer fetch from X."""
    if not MANUAL_OVERRIDES_PATH.is_file():
        return {}
    try:
        data = json.loads(MANUAL_OVERRIDES_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as e:
        print(f"  manual overrides unreadable: {e}", file=sys.stderr)
        return {}
    if not isinstance(data, dict):
        return {}
    out: dict[str, dict[str, str]] = {}
    for key, val in data.items():
        if isinstance(val, str) and val.strip():
            out[str(key)] = {"text": val.strip()}
        elif isinstance(val, dict):
            text = (val.get("text") or val.get("headline") or "").strip()
            if text:
                out[str(key)] = {
                    "text": text,
                    "created_at": (val.get("created_at") or "").strip(),
                }
    return out


def fetch_profile_bio(handle: str) -> str | None:
    api = FXTWITTER_USER.format(user=urllib.parse.quote(handle))
    req = urllib.request.Request(api, headers={"User-Agent": "cz-book-club-import/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError, OSError) as e:
        print(f"  profile bio fail: @{handle} ({e})", file=sys.stderr)
        return None
    user = data.get("user") if isinstance(data.get("user"), dict) else None
    if not user:
        return None
    bio = (user.get("description") or "").strip()
    return bio if bio and not is_weak_post_text(bio) else None


def import_stories(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    stories = []
    fetched = 0
    failed_fetch = 0
    manual_overrides = load_manual_overrides()

    for row in ws.iter_rows(min_row=1, max_row=120):
        name = (row[0].value or "").strip() if row[0].value else ""
        handle_raw = (row[1].value or "").strip() if row[1].value else ""
        reason_cell = row[2]
        reason = (reason_cell.value or "").strip() if reason_cell.value else ""
        hyperlink = reason_cell.hyperlink.target if reason_cell.hyperlink else None

        if not name or name in SKIP_NAMES or name.startswith("#"):
            continue
        if not handle_raw.startswith("@"):
            continue
        handle = handle_raw.lstrip("@").strip()
        if not handle:
            continue
        if should_exclude_handle(handle):
            print(f"  skip @{handle} (excluded)", file=sys.stderr)
            continue

        sheet_body = clean_sheet_reason(reason)
        url = (hyperlink or f"https://x.com/{handle}").strip()
        if url.startswith("http://"):
            url = "https://" + url[7:]
        url = url.replace("twitter.com", "x.com")

        body = sheet_body
        created_at = ""
        method = "kol_sheet_profile_fallback"
        score = 0.75

        if hyperlink and STATUS_RE.search(hyperlink):
            status_m = STATUS_RE.search(hyperlink)
            tweet_id = status_m.group(1) if status_m else ""
            user_from_url = HANDLE_FROM_URL_RE.search(hyperlink)
            screen_name = (user_from_url.group(1) if user_from_url else handle).lstrip("@")

            text, created = fetch_x_post_text(hyperlink)
            time.sleep(0.25)
            if text and is_weak_post_text(text):
                text = None
            if not text and tweet_id:
                text, created_vx = fetch_vxtwitter_post(screen_name, tweet_id)
                time.sleep(0.25)
                created = created or created_vx or ""
            if text:
                body = text
                created_at = created or ""
                method = "kol_sheet_x_live_text"
                score = 0.9
                fetched += 1
            elif tweet_id and tweet_id in manual_overrides:
                manual = manual_overrides[tweet_id]
                body = manual["text"]
                created_at = manual.get("created_at") or ""
                method = "kol_manual_archived_text"
                score = 0.88
                fetched += 1
                print(f"  manual text: @{screen_name}/{tweet_id}", file=sys.stderr)
            else:
                failed_fetch += 1
                method = "kol_sheet_reason_fallback_fetch_failed"
        elif hyperlink:
            # Telegram / other — keep sheet summary; link still points out
            method = "kol_sheet_external_link"
            score = 0.8
        elif normalize_handle(handle) in PROFILE_BIO_REFETCH_HANDLES:
            bio = fetch_profile_bio(handle)
            time.sleep(0.2)
            if bio:
                body = bio
                method = "kol_x_profile_bio"
                score = 0.85
                fetched += 1
            else:
                failed_fetch += 1
                method = "kol_sheet_profile_fallback_bio_failed"

        tweet_m = STATUS_RE.search(url)
        sid = (
            f"tw-{tweet_m.group(1)}"
            if tweet_m
            else "kol-" + hashlib.sha1(f"{handle}:{name}".encode()).hexdigest()[:12]
        )

        headline, display_body = finalize_headline_body(body)
        stories.append(
            {
                "id": sid,
                "kind": "twitter",
                "author_display": f"{name} · @{handle}",
                "headline": headline,
                "body": display_body,
                "url": normalize_x_url(url) if STATUS_RE.search(url) else url,
                "created_at": created_at,
                "author_avatar_url": f"https://unavatar.io/x/{handle}",
                "curator": {
                    "tier": "standard",
                    "score": score,
                    "method": method,
                },
            }
        )

    return {
        "format_version": 2,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "query": "local-xlsx + X oEmbed post text for status hyperlinks",
        "source_file": xlsx_path.name,
        "curator_pipeline": {
            "backend": "import_book_club_kol_stories.py",
            "heuristic": "reason_column_or_oembed_text",
            "openai": False,
            "exclude_official_accounts": False,
            "min_stories_target": len(stories),
            "oembed_fetched": fetched,
            "oembed_failed": failed_fetch,
            "live_text_note": "Status hyperlinks use X oEmbed + vxtwitter fallback (articles / link-only posts).",
        },
        "stories": stories,
    }


def main() -> int:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.is_file():
        print(f"Missing xlsx: {xlsx}", file=sys.stderr)
        return 1

    print(f"Reading {xlsx} …")
    doc = import_stories(xlsx)
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with OUT_JSON.open("w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)
        f.write("\n")

    pipe = doc["curator_pipeline"]
    print(
        f"Wrote {OUT_JSON} — {len(doc['stories'])} stories, "
        f"oEmbed OK: {pipe['oembed_fetched']}, failed: {pipe['oembed_failed']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
